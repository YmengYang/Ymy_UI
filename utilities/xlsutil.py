#!/usr/bin/env python3.8
# -*- coding: utf-8 -*-
# ---
# @Software: PyCharm
# @Site:
# @File: test_003_protocol_create.py
# @Author: PGQ
# @E-mail: panguoqing.pgq@sunyur.com
# @Time: 1月 30, 2021
# ---

from openpyxl import load_workbook


class XlsxUtil():

    def __init__(self, file, sheetname):
        self.filepath = file
        self.workbook = load_workbook(file)
        # self.workbook.active
        self.sheet_name = sheetname
        self.res = ""

    def set_sheet(self, sheetname):
        self.sheet_name = sheetname

    def get_max_rows(self):
        sheet = self.workbook[self.sheet_name]
        return sheet.max_row

    def get_max_cols(self):
        sheet = self.workbook[self.sheet_name]
        return sheet.max_column

    def get_cell_data(self, row, col):
        sheet = self.workbook[self.sheet_name]
        value = sheet.cell(row=row, column=col).value
        if value is None:
            return ""
        return value

    def set_cell_data(self, row, col, data):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row=row, column=col).value = data
        self.workbook.save(self.filepath)

    def get_title_col(self, title_name):
        """通过标题名title_name称获取列值"""
        sheet = self.workbook[self.sheet_name]
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(1, column=col).value
            if title_name == value:
                self.res = col
                break
        return self.res

    def close_wb(self):
        self.workbook.close()
